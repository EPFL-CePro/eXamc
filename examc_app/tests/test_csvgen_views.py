from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook


class CsvgenViewsTest(SimpleTestCase):
    def test_csvgen_template_uses_generated_template_download(self):
        template_path = Path(__file__).resolve().parents[2] / "templates" / "csvgen" / "csvgen.html"
        template_source = template_path.read_text()

        self.assertIn("{% url 'download_csvgen_templates' %}", template_source)
        self.assertNotIn("{% static 'templates.zip' %}", template_source)
        self.assertNotIn("images_csvgen/30.gif", template_source)

    def test_download_csvgen_templates_returns_expected_excel_files(self):
        response = self.client.get(reverse("download_csvgen_templates"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="csvgen_templates.zip"')

        with ZipFile(BytesIO(response.content)) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"amc_students_template.xlsx", "ans_students_template.xlsx"},
            )
            self.assert_excel_header(
                archive.read("amc_students_template.xlsx"),
                ["name", "sciper", "email", "seat", "room"],
            )
            self.assert_excel_header(
                archive.read("ans_students_template.xlsx"),
                ["class name", "email", "sciper", "name"],
            )

    def assert_excel_header(self, excel_content, expected_header):
        workbook = load_workbook(BytesIO(excel_content), read_only=True)
        worksheet = workbook.active

        self.assertEqual(
            [cell.value for cell in next(worksheet.iter_rows(max_row=1))],
            expected_header,
        )
